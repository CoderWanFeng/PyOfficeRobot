import pandas as pd
import uiautomation as auto
import time
import logging
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
# 配置日志


class WeChatBot:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    def __init__(self, file_name,contact_name):
        """
        初始化 WeChatBot 类，指定的联系人名称。
        """
        self.contact_name = contact_name
        self.wx = self.find_wechat_window()
        self.last_processed_msg = None
        self.file_name = file_name
        # 判断文件是否存在
        if os.path.isfile(self.file_name):
            logging.info(f"文件 '{self.file_name}' 已存在。")
        else:
            self.create_excel_file()

    def generate_date_range(self):
        """
        生成当前月份的日期范围
        """
        now = datetime.now()
        year = now.year
        month = now.month
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        last_day_of_month = next_month - timedelta(days=1)
        start_date = datetime(year, month, 1).strftime('%m月%d号')
        end_date = last_day_of_month.strftime('%m月%d号')
        return f"来客/视频会议信息({start_date}-{end_date})"

    def setup_worksheet(self,ws, title):
        """
        设置工作表的标题和表头
        """
        # 设置标题
        ws.merge_cells('A1:H1')  # 合并A1到H1单元格
        ws['A1'] = title  # 设置标题内容
        ws['A1'].font = Font(size=20)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 设置表头
        headers = ["日期时间", "客户", "来客/出席人数", "客户职位", "目的", "对应级别", "对应人员", "会议室"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
            ws.cell(row=2, column=col).font = Font(size=14)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')

    def set_column_widths(self,ws):
        """
        设置列宽
        """
        for col in range(1, 9):  # A到H列
            ws.column_dimensions[chr(64 + col)].width = 20

    def set_cell_styles(self, ws):    
        """
        设置单元格样式
        """
        max_row = ws.max_row  # 动态获取最大行
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def create_excel_file(self):
        """
        创建Excel文件
        """
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active

        # 生成标题
        title = self.generate_date_range()

        # 设置工作表
        self.setup_worksheet(ws, title)
        self.set_column_widths(ws)
        self.set_cell_styles(ws)

        # 保存文件
        wb.save(self.file_name)
        logging.info(f"Excel 文件 '{self.file_name}' 创建成功！")
        
    def find_wechat_window(self):
        """
        查找微信窗口并切换到该窗口。
        """
        try:
            wx_window = auto.WindowControl(searchDepth=3, Name='微信')
            if wx_window.Exists(0):
                wx_window.SwitchToThisWindow()
                return wx_window
            else:
                logging.warning("未找到微信窗口")
                return None
        except Exception as e:
            logging.error(f"查找微信窗口出错: {e}")
            return None

    def select_conversation(self):
        """
        选择指定的联系人对话。
        """
        conversations = self.wx.ListControl()
        for conversation in conversations.GetChildren():
            if conversation.Name == self.contact_name:
                conversation.Click(simulateMove=False)
                break

    def get_last_message(self):
        """
        获取对话中的最后一条消息。
        """
        try:
            message_list = self.wx.ListControl(Name='消息').GetChildren()
            if message_list:
                last_msg = message_list[-1].Name
                logging.info(f"获取到最后一条消息: {last_msg}")
                return last_msg
            else:
                logging.info("消息列表为空")
                return None
        except Exception as e:
            logging.error(f"获取最后一条消息出错: {e}")
            return None




    def process_messages(self):
        """
        处理新消息并保存到 Excel 文件。
        """
        while True:
            time.sleep(3)
            last_msg = self.get_last_message()
            if last_msg and last_msg != self.last_processed_msg:
                if "会议信息" in last_msg:
                    # 解析信息
                    message = last_msg
                    message = message.strip()
                    lines = message.split('\n')
                    meeting_info_list = []
                    for line in lines:
                        if '：' in line:
                            _, value = line.split('：', 1)
                            meeting_info_list.append(value.strip())
                    data = meeting_info_list
                    # 将数据写入 Excel 文件
                    file_path = self.file_name
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active
                    # 查找最后一行的有效数据
                    next_row = sheet.max_row + 1
                    if next_row == 1 and not sheet.cell(row=1, column=1).value:
                        next_row = 1  # 如果第一行是空的，设置为1
                    for col, value in enumerate(data, start=1):
                        sheet.cell(row=next_row, column=col, value=value)
                    try:
                        workbook.save(file_path)
                        logging.info(f"数据已成功写入到 {file_path}")
                    except Exception as e:
                        logging.info(f"保存文件时出错: {e}")
                    self.last_processed_msg = last_msg
                    logging.info(f"消息已保存至{self.file_name}")
                else:
                    logging.info("消息不包含会议信息关键字，跳过处理")
            else:
                logging.info("没有新消息或消息已处理")

# 示例调用-需要创建的文件名，联系人名称
bot = WeChatBot('InfoData.xlsx','测试群2')
if bot.wx:
    bot.select_conversation()
    bot.process_messages()



