# -*- codeing = utf-8 -*-
# @Time : 2023/3/14 18:18
# @Project_File : Start.py
# @Dir_Path : _____Release_File/微信自动化/微信群发消息
# @File : QT.py
# @IDE_Name : PyCharm 
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================

import sys
import threading

from openpyxl import load_workbook
from PySide6.QtCore import QStringListModel, QModelIndex
from PySide6.QtWidgets import QApplication, QWidget
from PySide6.QtWidgets import QFileDialog

from PyOfficeRobot.api.chat import send_message
from PyOfficeRobot.core.group.ui_file_py import Ui_Form


# 继承QWidget类，以获取其属性和方法
class MyWidget(QWidget):
    def __init__(self):
        super().__init__()
        # 设置界面为我们生成的界面
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.set_ui()

        self.sendtext: str = str()
        self.groupslist: list = list()
        self.groupslistfilename: str = str()
        self.sendtextfilename: str = str()
        self.state = True

    def set_ui(self):
        self.setWindowTitle("Python自动化办公社区")
        self.ui.pushButton_load.clicked.connect(self.OnPushbutton_load_Cliked)
        self.ui.pushButton_start.clicked.connect(self.OnPushbutton_Start_Cliked)
        self.ui.pushButton_close.clicked.connect(self._OnPushbutton_Close_Clicked)
        self.ui.listView.clicked.connect(self.test)

    def test(self, index: QModelIndex):
        count = self.model.rowCount()
        print(count)
        for _ in range(count):
            print(self.model.index(_).data())

    def OnPushbutton_load_Cliked(self):
        self.groupslist.clear()

        self.groupslistfilename = QFileDialog.getOpenFileName(caption='选择群发列表', filter='*.xlsx')[0]
        self.sendtextfilename = QFileDialog.getOpenFileName(caption='选择群发的文本', filter='*.txt')[0]

        # <editor-fold desc="代码段 ： 设置群发好友列表">
        wb = load_workbook(filename=self.groupslistfilename)
        sheet = wb.active
        rows = sheet.max_row

        for _ in range(1, rows + 1):
            itme = sheet.cell(row=_, column=1).value
            state = sheet.cell(row=_, column=2).value

            if str(state) == "True":
                if itme:
                    self.groupslist.append(str(itme))
        # </editor-fold>

        # <editor-fold desc="代码段 ：设置群发文本">
        with open(file=self.sendtextfilename, encoding='utf-8') as file:
            self.sendtext = '{ctrl}{ENTER}'.join(file.readlines())
        # </editor-fold>

        # <editor-fold desc="代码段 : 设置群发列表到ui">
        self.model = QStringListModel()
        self.model.setStringList(self.groupslist)
        self.ui.listView.setModel(self.model)
        # </editor-fold>

    def threadfunction(self):
        for _ in self.groupslist:
            thread_start = threading.Thread(target=send_message, args=(_, self.sendtext))
            thread_start.start()
            thread_start.join()

    def OnPushbutton_Start_Cliked(self):
        self.thread_1 = threading.Thread(target=self.threadfunction)
        self.thread_1.start()

    def _OnPushbutton_Close_Clicked(self):
        sys.exit(0)


# 程序入口
if __name__ == "__main__":
    app = QApplication(sys.argv)
    # 初始化QApplication，界面展示要包含在QApplication初始化之后，结束之前

    window = MyWidget()
    window.show()
    # 初始化并展示我们的界面组件

    sys.exit(app.exec_())
    # 结束QApplication
