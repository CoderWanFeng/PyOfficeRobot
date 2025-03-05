# -*- coding: UTF-8 -*-
'''
@作者 ：B站/抖音/微博/小红书/公众号，都叫：程序员晚枫
@读者群     ：http://www.python4office.cn/wechat-group/
@个人网站 ：www.python-office.com
@Date    ：2023/4/27 21:50 
@Description     ：
'''

import sys

from PyOfficeRobot.lib.decorator_utils.instruction_url import instruction
from PySide6.QtWidgets import QApplication

from PyOfficeRobot.core.group.Start import MyWidget

@instruction
def send():
    app = QApplication(sys.argv)
    # 初始化QApplication，界面展示要包含在QApplication初始化之后，结束之前
    window = MyWidget()
    window.show()
    # 初始化并展示我们的界面组件
    sys.exit(app.exec_())
    # 结束QApplication


import pandas as pd
import uiautomation as auto
import time
import logging

class WeChatBot:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    def __init__(self, csv_path, contact_name):
        self.df = pd.read_csv(csv_path)
        self.contact_name = contact_name
        self.wx = self.find_wechat_window()
        self.last_processed_msg = None

    def find_wechat_window(self):
        try:
            wx_window = auto.WindowControl(searchDepth=3, Name='微信')
            if wx_window.Exists(0):
                wx_window.SwitchToThisWindow()
                return wx_window
            else:
                logging.warning("未找到微信窗口")
                return None
        except Exception as e:
            logging.error(f"查找微信窗口出错: {e}")
            return None

    def select_conversation(self):
        conversations = self.wx.ListControl()
        for conversation in conversations.GetChildren():
            if conversation.Name == self.contact_name:
                conversation.Click(simulateMove=False)
                break

    def get_last_message(self):
        try:
            message_list = self.wx.ListControl(Name='消息').GetChildren()
            if message_list:
                last_msg = message_list[-1].Name
                logging.info(f"获取到最后一条消息: {last_msg}")
                return last_msg
            else:
                logging.info("消息列表为空")
                return None
        except Exception as e:
            logging.error(f"获取最后一条消息出错: {e}")
            return None

    def send_reply(self, reply):
        try:
            self.wx.SendKeys(reply, waitTime=0)
            self.wx.SendKeys('{Enter}', waitTime=1)
            logging.info(f"回复内容是: {reply}")
        except Exception as e:
            logging.error(f"发送消息出错: {e}")

    def process_messages(self):
        while True:
            time.sleep(3)
            last_msg = self.get_last_message()
            if last_msg and last_msg != self.last_processed_msg:
                if last_msg not in self.df['回复内容'].values:
                    matched_replies = self.df[self.df['关键词'].apply(lambda x: x in last_msg)]['回复内容']
                    if not matched_replies.empty:
                        for reply in matched_replies:
                            reply = reply.replace('{br}', '\n')
                            self.send_reply(reply)
                        self.last_processed_msg = last_msg
                    else:
                        logging.info("没有匹配的关键字")
                else:
                    logging.info("最后一条消息是自动回复内容，跳过回复")
            else:
                logging.info("没有新消息或消息已处理")


import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta

class WeChatBot2:
    # logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    def __init__(self, file_name,contact_name):
        """
        初始化 WeChatBot 类，指定的联系人名称。
        """
        self.contact_name = contact_name
        self.wx = self.find_wechat_window()
        self.last_processed_msg = None
        self.file_name = file_name
        # 判断文件是否存在
        if os.path.isfile(self.file_name):
            logging.info(f"文件 '{self.file_name}' 已存在。")
        else:
            self.create_excel_file()

    def generate_date_range(self):
        """
        生成当前月份的日期范围
        """
        now = datetime.now()
        year = now.year
        month = now.month
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        last_day_of_month = next_month - timedelta(days=1)
        start_date = datetime(year, month, 1).strftime('%m月%d号')
        end_date = last_day_of_month.strftime('%m月%d号')
        return f"来客/视频会议信息({start_date}-{end_date})"

    def setup_worksheet(self,ws, title):
        """
        设置工作表的标题和表头
        """
        # 设置标题
        ws.merge_cells('A1:H1')  # 合并A1到H1单元格
        ws['A1'] = title  # 设置标题内容
        ws['A1'].font = Font(size=20)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 设置表头
        headers = ["日期时间", "客户", "来客/出席人数", "客户职位", "目的", "对应级别", "对应人员", "会议室"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
            ws.cell(row=2, column=col).font = Font(size=14)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')

    def set_column_widths(self,ws):
        """
        设置列宽
        """
        for col in range(1, 9):  # A到H列
            ws.column_dimensions[chr(64 + col)].width = 20

    def set_cell_styles(self, ws):
        """
        设置单元格样式
        """
        max_row = ws.max_row  # 动态获取最大行
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=8):
            for cell in row:
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def create_excel_file(self):
        """
        创建Excel文件
        """
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active

        # 生成标题
        title = self.generate_date_range()

        # 设置工作表
        self.setup_worksheet(ws, title)
        self.set_column_widths(ws)
        self.set_cell_styles(ws)

        # 保存文件
        wb.save(self.file_name)
        logging.info(f"Excel 文件 '{self.file_name}' 创建成功！")

    def find_wechat_window(self):
        """
        查找微信窗口并切换到该窗口。
        """
        try:
            wx_window = auto.WindowControl(searchDepth=3, Name='微信')
            if wx_window.Exists(0):
                wx_window.SwitchToThisWindow()
                return wx_window
            else:
                logging.warning("未找到微信窗口")
                return None
        except Exception as e:
            logging.error(f"查找微信窗口出错: {e}")
            return None

    def select_conversation(self):
        """
        选择指定的联系人对话。
        """
        conversations = self.wx.ListControl()
        for conversation in conversations.GetChildren():
            if conversation.Name == self.contact_name:
                conversation.Click(simulateMove=False)
                break

    def get_last_message(self):
        """
        获取对话中的最后一条消息。
        """
        try:
            message_list = self.wx.ListControl(Name='消息').GetChildren()
            if message_list:
                last_msg = message_list[-1].Name
                logging.info(f"获取到最后一条消息: {last_msg}")
                return last_msg
            else:
                logging.info("消息列表为空")
                return None
        except Exception as e:
            logging.error(f"获取最后一条消息出错: {e}")
            return None




    def process_messages(self):
        """
        处理新消息并保存到 Excel 文件。
        """
        while True:
            time.sleep(3)
            last_msg = self.get_last_message()
            if last_msg and last_msg != self.last_processed_msg:
                if "会议信息" in last_msg:
                    # 解析信息
                    message = last_msg
                    message = message.strip()
                    lines = message.split('\n')
                    meeting_info_list = []
                    for line in lines:
                        if '：' in line:
                            _, value = line.split('：', 1)
                            meeting_info_list.append(value.strip())
                    data = meeting_info_list
                    # 将数据写入 Excel 文件
                    file_path = self.file_name
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active
                    # 查找最后一行的有效数据
                    next_row = sheet.max_row + 1
                    if next_row == 1 and not sheet.cell(row=1, column=1).value:
                        next_row = 1  # 如果第一行是空的，设置为1
                    for col, value in enumerate(data, start=1):
                        sheet.cell(row=next_row, column=col, value=value)
                    try:
                        workbook.save(file_path)
                        logging.info(f"数据已成功写入到 {file_path}")
                    except Exception as e:
                        logging.info(f"保存文件时出错: {e}")
                    self.last_processed_msg = last_msg
                    logging.info(f"消息已保存至{self.file_name}")
                else:
                    logging.info("消息不包含会议信息关键字，跳过处理")
            else:
                logging.info("没有新消息或消息已处理")